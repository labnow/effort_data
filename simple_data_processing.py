# -*- coding: utf-8 -*-

"""
@File        : simple_data_processing.py
@Author      :
@Time        : 2022/2/21 15:00
@Description : processing project data
"""

# data source: 'data4database/revenue.xlsx', 'data4database/internal_cost.xlsx', 'data4database/external_cost.xlsx'
# output: 'data4pivot/dataDumped.csv', 'data4pivot/dataExpense.csv', 'data4pivot/dataProfit.csv', 'data4pivot/dataRevenue.csv' 

t_config = {
    "revenue": {
        "filePath": "data4database/revenue.xlsx",
        "sheetName": "default",
        "dataStartFromRow": 4,
        "dataEndAtRow": 54,
        "dataStartFromColumn": 3,
        "dataEndAtColumn": 5,
        "tableHeaderRow": 3,
        "jsonPath": 'data4pivot/revenue.json'
    },
    "internal_cost": {
        "filePath": "data4database/internal_cost.xlsx",
        "sheetName": "default",
        "dataStartFromRow": 4,
        "dataEndAtRow": 80,
        "dataStartFromColumn": 3,
        "dataEndAtColumn": 5,
        "tableHeaderRow": 3,
        "jsonPath": 'data4pivot/internal_cost.json'
    },
    "external_cost": {
        "filePath": "data4database/external_cost.xlsx",
        "sheetName": "default",
        "dataStartFromRow": 4,
        "dataEndAtRow": 633,
        "dataStartFromColumn": 3,
        "dataEndAtColumn": 8,
        "tableHeaderRow": 3,
        "jsonPath": 'data4pivot/external_cost.json'
    }
}

def xlsx2json(t_config):
    from openpyxl import load_workbook
    import json

    t_meta = t_config # metadata dict
    t_wb = load_workbook(t_meta['filePath'], read_only=True, data_only=True)
    t_ws = t_wb[t_meta['sheetName']]
    
    col_start = t_meta['dataStartFromColumn']
    col_end = t_meta['dataEndAtColumn']
    row_start = t_meta['dataStartFromRow']
    row_end = t_meta['dataEndAtRow']
    header_row = t_meta['tableHeaderRow']
    current_row = 1
    table_header = []
    data_dict = {}
    # get table header dict
    for col in range(col_start, col_end + 1):
        table_header.append(t_ws.cell(row=header_row, column=col).value)

    index_of_date_col = table_header.index('date')
    # get table header dict
    for row in t_ws.iter_rows(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end, values_only=True):
        row = list(row)
        try:
            row[index_of_date_col] = str(row[index_of_date_col].date()) # json encode str(datetime)
        except:
            pass

        data_dict[current_row] = row
        current_row += 1

    all_dict = {'metadata':t_meta, 'header': table_header, 'data': data_dict}
    with open(t_meta['jsonPath'], 'w') as json_file:
        json.dump(all_dict, json_file, ensure_ascii=False, indent=4)
        
    return current_row

def json2csv():
    import json
    import csv

    columns_data_dumped= ['project','amount', 'invoiceDate', 'type']
    columns_data_expense = ['project','billDate','expense','category', 'partner']
    t_data_dumped = list()
    t_data_expense = list()
    with open('data4pivot/revenue.json', 'r') as f:
        revenue = json.load(f)['data']

    with open('data4pivot/external_cost.json', 'r') as f:
        external_cost = json.load(f)['data']

    with open('data4pivot/internal_cost.json', 'r') as f:
        intenal_cost = json.load(f)['data']

    for key, item in revenue.items():
        t_data_dumped.append((item[0], item[2]*1.06, item[1], '03-Revenue'))

    for key, item in external_cost.items():
        t_data_dumped.append((item[0], item[3]*-1, item[2], '02-RBEI'))
        t_data_expense.append((item[0], item[2], item[3], item[4], item[5]))

    for key, item in intenal_cost.items():
        t_data_dumped.append((item[0], item[2]*-1, item[1], '01-Vendor'))
    
    with open('data4pivot/dataDumped.csv', 'w', encoding='UTF8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(columns_data_dumped)
        writer.writerows(t_data_dumped)

    with open('data4pivot/dataExpense.csv', 'w', encoding='UTF8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(columns_data_expense)
        writer.writerows(t_data_expense)

def run():
    xlsx2json(t_config["revenue"])
    xlsx2json(t_config["internal_cost"])
    xlsx2json(t_config["external_cost"])
    json2csv()

if __name__ == '__main__':
    run()
    print('OK')
