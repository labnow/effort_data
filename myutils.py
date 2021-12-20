from datetime import datetime, date
import json
from openpyxl import load_workbook
import sys
# from app import db, ExpenseInternal, Revenue, ExpenseExternal
import os
from log_handler import logger

def process_reports_from_vendor(folder_to_vendor_reports):
    t_config = read_config_json()
    logger.info('Start reading all reports in this {}'.format(folder_to_vendor_reports))
    reports_to_process = list()
    for root, dirs, files in os.walk(folder_to_vendor_reports, topdown=True):
        for name in files:
            if name.split('_')[0] in t_config['vendor_whitelist'] and name.split('.')[-1] == 'xlsx' and name not in t_config['vendor_reports']['processed_reports']:
                reports_to_process.append(name)
                # logger.info(os.path.join(root, name))
                logger.info('Processing {}'.format(name))
        
    return reports_to_process

def db2xlsx():
    wb = load_workbook('data_from_vendor.xlsx')
    ws = wb['default']

    external_expense = ExpenseExternal.query.order_by(ExpenseExternal.invoice_date)
    row_num = 4
    for item in external_expense:
        ws.cell(row=row_num, column=3, value=item.project)
        ws.cell(row=row_num, column=4, value=item.item)
        ws.cell(row=row_num, column=5, value=item.invoice_date)
        ws.cell(row=row_num, column=6, value=item.amount)
        ws.cell(row=row_num, column=7, value=item.category)
        ws.cell(row=row_num, column=8, value=item.partner)
        row_num += 1

    wb.save('data_from_vendor.xlsx')

def audi2excel():
    t_dict = []
    # t_dict = t_dict + audi2python(4) + audi2python(10) + audi2python(16) + audi2python(22) + audi2python(28) + audi2python(34)
    t_dict = t_dict + audi2python(16)
    wb = load_workbook('data_from_vendor_audi.xlsx')
    ws = wb['default']

    row_num = 490
    for item in t_dict:
        # t_dict = {'week':week, 'who':who, 'date':date, 'starttime':starttime, 'endtime':endtime, 'workinghours':workinghours, 'overtime':overtime, 'location':location, 'worklog':worklog}
        ws.cell(row=row_num, column=3, value=item['who'])
        ws.cell(row=row_num, column=4, value=item['week'])
        ws.cell(row=row_num, column=5, value=item['date'])
        ws.cell(row=row_num, column=6, value=item['starttime'])
        ws.cell(row=row_num, column=7, value=item['endtime'])
        ws.cell(row=row_num, column=8, value=item['workinghours'])
        ws.cell(row=row_num, column=9, value=item['overtime'])
        ws.cell(row=row_num, column=10, value=item['location'])
        ws.cell(row=row_num, column=11, value=item['worklog'])
        row_num += 1

    wb.save('data_from_vendor_audi.xlsx')

def audi2python(r_in):
    wb = load_workbook('guoqi.xlsx', data_only=True)
    ws = wb['default']

    row_num = 0
    col_num = 0
    
    all_dict = []

    for c in range(3, 112):
        # for r in range(1, 7):
        r = r_in
        starttime = ws.cell(row=r, column=c).value
        if not starttime or starttime == '00:00:00' or starttime == '0:00:00' or starttime == '-':
            continue
        r += 1
        endtime = ws.cell(row=r, column=c).value
        r += 1
        workinghours = ws.cell(row=r, column=c).value
        r += 1
        overtime = ws.cell(row=r, column=c).value
        r += 1
        location = ws.cell(row=r, column=c).value
        r += 1
        worklog = ws.cell(row=r, column=c).value
        week = ws.cell(row=1, column=c).value
        date = ws.cell(row=2, column=c).value
        who = ws.cell(row=r, column=1).value
        t_dict = {'week':week, 'who':who, 'date':date, 'starttime':starttime, 'endtime':endtime, 'workinghours':workinghours, 'overtime':overtime, 'location':location, 'worklog':worklog}
        all_dict.append(t_dict)
        # print(t_dict)
    
    print(len(all_dict))
    return all_dict

# JSON file
def read_config_json():
    with open('config/config.json', "r") as f:
        myconfig = json.load(f)

    if myconfig:
        # print('config load successfully.')
        return myconfig
    else:
        print('config not exist<br>')

def check_data(my_project, my_date, my_amount):
    myconfig = read_config_json()
    
    if not my_project in myconfig['project_whitelist']:
        logger.error('<b>!!!Invalid Project Name</b><br>')
        return False
    if not isinstance(my_date, date):
        logger.error('<b>!!!Invalid Invoice Date</b><br>')
        return False
    if not isinstance(my_amount, (int, float, complex)):
        logger.error('<b>!!!Invalid Amount</b><br>')
        return False

    return True

def read_external():
    t_meta = read_config_json()['external_report']

    # metadata dict
    t_wb = load_workbook(t_meta['filePath'], read_only=True, data_only=True)
    t_ws = t_wb[t_meta['sheetName']]
    
    col_start = t_meta['dataStartFromColumn']
    col_end = t_meta['dataEndAtColumn']
    row_start = t_meta['dataStartFromRow']
    row_end = t_meta['dataEndAtRow']
    header_row = t_meta['headerRow']
    current_row = 1
    table_header = []
    data_dict = {}
    # get table header dict
    for col in range(col_start, col_end + 1):
        table_header.append(t_ws.cell(row=header_row, column=col).value)

    index_of_date_col = table_header.index('date')
    # get table header dict
    for row in t_ws.iter_rows(min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end, values_only=True):
        row = list(row)
        row[index_of_date_col] = str(row[index_of_date_col].date()) # json encode str(datetime)
        data_dict[current_row] = row
        current_row += 1

    all_dict = {'metadata':t_meta, 'header': table_header, 'data': data_dict}
    with open(t_meta['jsonPath'], 'w') as json_file:
        json.dump(all_dict, json_file, ensure_ascii=False, indent=4)
        
    return 'OK'

# def read_external():
#     myconfig = read_config_json()['external_effort']
#     return excel_to_json(myconfig['filePath'], myconfig['sheetName'], myconfig['headerRow'], myconfig['dataStartFromRow'], myconfig['dataEndAtRow'], myconfig['dataStartFromColumn'], myconfig['dataEndAtColumn'], myconfig['jsonPath'])

def process_ecncost():
    logger.info('<h2>Reading ecn_cost/internal_expense data...</h2>')
    myconfig = read_config_json()
    wb = load_workbook(myconfig['ecnCost']['filePath'], data_only=True)
    ws = wb[myconfig['ecnCost']['sheetName']]

    myData = []
    myDataTotal = 0

    row_start = myconfig['ecnCost']['dataStartFromRow']
    row_end = myconfig['ecnCost']['dataEndAtRow']
    column_start = myconfig['ecnCost']['dataStartFromColumn'] # B
    column_end = myconfig['ecnCost']['dataEndAtColumn'] # K
    table_header = myconfig['ecnCost']['tableHeaderRow']
    table_index = myconfig['ecnCost']['tableIndexColumn']
    t_header = ['project', 'month', 'amount']
    t_metadata = myconfig['ecnCost']
    data_dict = dict()
    record_num = 1

    for row in range(row_start, row_end + 1):
        for col in range(column_start, column_end + 1):
            tmp_cell = ws.cell(row=row, column=col)
            if tmp_cell.value:
                myProject = ws.cell(row=table_header, column=col).value             
                myDate = date(2021,ws.cell(row=row, column=table_index).value,1)
                myAmount = tmp_cell.value
                if check_data(myProject, myDate, myAmount):
                    myDataTotal += myAmount
                    myTuple = (myProject, str(myDate), myAmount)
                    data_dict[record_num] = myTuple
                    record_num += 1
                    myData.append(myTuple)
                    logger.info('{} add to database successfully<br>'.format(myTuple))
                else:
                    logger.error('{} {} {} has been dropped<br>'.format(myProject, str(myDate), myAmount))

    all_dict = {'metadata':t_metadata, 'header': t_header, 'data': data_dict}
    json_file_name = myconfig['ecnCost']['jsonPath']
    with open(json_file_name, 'w') as json_file:
            json.dump(all_dict, json_file, ensure_ascii=False, indent=4)
    myDataLength = len(myData)
    logger.info('<h3>{} value read, sum of amount is {}</h3>'.format(str(myDataLength), str(myDataTotal)))

def process_billing_status():
    logger.info('<h2>Reading billing_status/revenue data...</h2>')
    myconfig = read_config_json()

    wb = load_workbook(myconfig['billing_status']['filePath'], data_only=True)
    ws = wb[myconfig['billing_status']['sheetName']]

    row_start = myconfig['billing_status']['dataStartFromRow']
    row_end = myconfig['billing_status']['dataEndAtRow']
    column_project = myconfig['billing_status']['column_project']
    column_date = myconfig['billing_status']['column_date']
    column_amount = myconfig['billing_status']['column_amount']
    t_header = ['project', 'invoice_date', 'amount']
    t_metadata = myconfig['billing_status']

    myData = []
    myDataTotal = 0
    data_dict = dict()
    record_num = 1

    for row in range(row_start, row_end + 1):
        # read property:date
        myDate = ws.cell(row=row, column=column_date).value
        # read property:project
        row_tmp = row
        while True:
            myProject = ws.cell(row=row_tmp, column=column_project).value
            if myProject:
                break
            row_tmp -= 1
        # read property:amount
        myAmount = ws.cell(row=row, column=column_amount).value
        if check_data(myProject, myDate, myAmount):
            myDataTotal += myAmount
            myTuple = (myProject, str(myDate.date()), myAmount)
            data_dict[record_num] = myTuple
            record_num += 1
            myData.append(myTuple)
            logger.info('{} add to database successfully<br>'.format(myTuple))
        else:
            logger.error('{} {} {} has been dropped<br>'.format(myProject, str(myDate), myAmount))

    all_dict = {'metadata':t_metadata, 'header': t_header, 'data': data_dict}
    json_file_name = myconfig['billing_status']['jsonPath']
    with open(json_file_name, 'w') as json_file:
            json.dump(all_dict, json_file, ensure_ascii=False, indent=4)

    myDataLength = len(myData)
    logger.info('<h3>{} value read, sum of amount is {}</h3>'.format(str(myDataLength), str(myDataTotal)))

def read_external_effort():
    import shutil
    print('<h2>Reading external_efoort...</h2>')
    myconfig = read_config_json()

    try:
        wb = load_workbook(myconfig['external_effort']['filePath'], data_only=True)
    except:
        print('<h3>externalExpense.xlsx does not exist, please check!</h3>')
        return 0
    
    ws = wb.active

    myData = []
    myDataTotal = 0
    for row in ws.values:
        # excel header ['category', 'project', 'item', 'amount', 'invoice_date', 'partner']
        current_data = [x for x in row]
        if check_data(current_data[1], current_data[4], current_data[3]):
            try:
                expense_external = ExpenseExternal(project=current_data[1], item=current_data[2], invoice_date=current_data[4], amount=current_data[3], category=current_data[0], partner=current_data[5])
                db.session.add(expense_external)
                db.session.commit()
                myDataTotal += current_data[3]
                myData.append(current_data)
                print('{} add to database successfully<br>'.format(current_data))
            except:
                print('{} can not be added to database<br>'.format(current_data))         
        else:
            print('{} has been dropped<br>'.format(current_data))

    now = str(datetime.now())[:10]
    os.rename(myconfig['external_effort']['filePath'], myconfig['external_effort']['filePath'] + '_' + now)
    shutil.copyfile('data4database/externalExpense_Template.xlsx', 'data4database/externalExpense.xlsx')
    myDataLength = len(myData)
    print('<h3>{} value read, sum of amount is {}</h3>'.format(str(myDataLength), str(myDataTotal)))

def stage1_data2json():
    print('<h1>Stage1: Read data into JSON database</h1>')
    # myconfig = read_config_json()
    read_ecncost()
    process_billing_status()
    read_external_effort()

def stage2_db2csv():
    logger.info('<h1>Stage2: Export data to csv</h1>')
    import pandas as pd
    import sqlite3
    myconfig = read_config_json()

    conn = sqlite3.connect('effort_data.db')
    cur = conn.cursor()

    # refresh dataProfit.csv
    cols1 = ['project','revenue_total','expense_total','profit']
    cur.execute('SELECT project, sum(revenue), sum(expense), sum(revenue) - sum(expense) FROM (SELECT project AS project, amount AS expense, 0 AS revenue FROM expense_external UNION SELECT project AS project, 0 AS expense, amount*1.06 AS revenue FROM revenue AS foo) GROUP BY project')
    result1 = pd.DataFrame(cur, columns=cols1)
    file_to_write1 = myconfig['path_to_data4pivot'] + '/dataProfit.csv'
    result1.to_csv(file_to_write1,index=False)

    # refresh dataExpense.csv
    cols2 = ['project','billDate','expense','category', 'partner']
    cur.execute('SELECT project, invoice_date, amount, category, partner FROM expense_external')
    result2 = pd.DataFrame(cur, columns=cols2)
    file_to_write2 = myconfig['path_to_data4pivot'] + '/dataExpense.csv'
    result2.to_csv(file_to_write2,index=False)

    # refresh dataRevenue.csv
    cols3 = ['project','billDate','revenue']
    cur.execute('SELECT project, invoice_date, amount FROM revenue')
    result3 = pd.DataFrame(cur, columns=cols3)
    file_to_write3 = myconfig['path_to_data4pivot'] + '/dataRevenue.csv'
    result3.to_csv(file_to_write3,index=False)

    # refresh dataDumped.csv
    cols4 = ['project','amount', 'billDate', 'type']
    cur.execute('''SELECT project, amount, billDate, type FROM (SELECT project AS project, amount*(-1) AS amount, invoice_date AS billDate, '01-Vendor' AS type FROM expense_external UNION SELECT project AS project, amount*(-1) AS amount, invoice_date AS billDate, '02-RBEI' AS type FROM expense_internal UNION SELECT project AS project, amount*1.06 AS amount, invoice_date AS billDate, '03-Revenue' AS type FROM revenue AS foo)''')
    result4 = pd.DataFrame(cur, columns=cols4)
    file_to_write4 = myconfig['path_to_data4pivot'] + '/dataDumped.csv'
    result4.to_csv(file_to_write4, index=False)

    conn.close()

def data2csv():
    import csv

    logger.info('<h1>Stage2: Export data to csv</h1>')
    columns_data_dumped= ['project','amount', 'invoiceDate', 'type']
    columns_data_expense = ['project','billDate','expense','category', 'partner']
    t_data_dumped = list()
    t_data_expense = list()
    with open('billing_status.json', 'r') as f:
        revenue = json.load(f)['data']

    with open('from_vendor.json', 'r') as f:
        expense = json.load(f)['data']

    with open('ecn_cost.json', 'r') as f:
        rbei = json.load(f)['data']

    for key, item in revenue.items():
        t_data_dumped.append((item[0], item[2]*1.06, item[1], '03-Revenue'))

    for key, item in expense.items():
        t_data_dumped.append((item[0], item[3]*-1, item[2], '02-RBEI'))
        t_data_expense.append((item[0], item[2], item[3], item[4], item[5]))

    for key, item in rbei.items():
        t_data_dumped.append((item[0], item[2]*-1, item[1], '01-Vendor'))
    
    with open('dataDumped.csv', 'w', encoding='UTF8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(columns_data_dumped)
        writer.writerows(t_data_dumped)

    with open('dataExpensecsv', 'w', encoding='UTF8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(columns_data_expense)
        writer.writerows(t_data_expense)

if __name__ == '__main__':
    print('OK')
    # db2xlsx()
    # t_all_dict = []
    # t_all2 = t_all_dict + audi2python(4) + audi2python(10) + audi2python(16) + audi2python(22) + audi2python(28) + audi2python(34)

    # print(len(t_all2))
    # audi2excel()
    # read_ecncost()
    # process_billing_status()
    # read_external()
    # read_vendor_reports('reportFromVendor')